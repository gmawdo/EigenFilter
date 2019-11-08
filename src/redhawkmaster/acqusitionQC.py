import os
import sys

from laspy.file import File
import numpy as np
import time
import csv
import datetime
import re
from collections import defaultdict
import bisect
from datetime import datetime, timedelta
from gnsscal import date2gpswd
from sklearn.neighbors import NearestNeighbors
import math
import cv2
from skimage import io
import matplotlib.pyplot as plt
import shapefile
from pyproj import Proj, transform
from shapely.geometry import mapping, Polygon, Point, shape
import fiona
import xlsxwriter
from openpyxl import load_workbook

# https://stackoverflow.com/questions/33415475/how-to-get-current-date-and-time-from-gps-unsegment-time-in-python
# Needs to be updated in some time
_LEAP_DATES = ((1981, 6, 30), (1982, 6, 30), (1983, 6, 30),
               (1985, 6, 30), (1987, 12, 31), (1989, 12, 31),
               (1990, 12, 31), (1992, 6, 30), (1993, 6, 30),
               (1994, 6, 30), (1995, 12, 31), (1997, 6, 30),
               (1998, 12, 31), (2005, 12, 31), (2008, 12, 31),
               (2012, 6, 30), (2015, 6, 30), (2016, 12, 31))

LEAP_DATES = tuple(datetime(i[0], i[1], i[2], 23, 59, 59) for i in _LEAP_DATES)


def leap(date):
    """
    Return the number of leap seconds since 1980-01-01

    :param date: datetime instance
    :return: leap seconds for the date (int)
    """
    # bisect.bisect returns the index `date` would have to be
    # inserted to keep `LEAP_DATES` sorted, so is the number of
    # values in `LEAP_DATES` that are less than `date`, or the
    # number of leap seconds.
    return bisect.bisect(LEAP_DATES, date)


def gps2utc(week, secs):
    """
    Convert the gps week and sec into utc date.

    :param week: GPS week number, i.e. 1866
    :param secs: number of seconds since the beginning of `week`
    :return: datetime instance with UTC time
    """
    secs_in_week = 604800
    gps_epoch = datetime(1980, 1, 6, 0, 0, 0)
    date_before_leaps = gps_epoch + timedelta(seconds=week * secs_in_week + secs)
    return date_before_leaps - timedelta(seconds=leap(date_before_leaps))


def attribute(test, las_file, csv_name='attribute.csv'):
    """
    Extract the attribute Max Min and Range information from the file and write it in a CSV file.

    :param test: laspy object from which we get all attributes and output Max,Min and Range in a CSV file
    :type test: Laspy Object
    :param las_file: name of the laspy_object, i.e. 'Test.las'
    :type las_file: string
    :param csv_name: Name of the csv file
    :type csv_name: string
    :return  create or appends to csv file that csv_name as name
    """

    # All local attributes that are in Laspy object
    attr_keys = ['File Name', 'Max X', 'Min X', 'Range X', 'Max Y', 'Min Y', 'Range Y', 'Max Z', 'Min Z', 'Range Z',
                 'Max Intensity', 'Min Intensity', 'Range Intensity',
                 'Max Return Number', 'Min Return Number', 'Range Return Number',
                 'Max Number of Returns', 'Min Number of Returns', 'Range Number of Returns',
                 'Max Gps Time', 'Min Gps Time', 'Range Gps Time',
                 'Max Date Time', 'Min Date Time', 'Range Date Time',
                 'Max Scan Direction Flag', 'Min Scan Direction Flag', 'Range Scan Direction Flag',
                 'Max Edge of Flight Line', 'Min Edge of Flight Line', 'Range Edge of Flight Line',
                 'Max Classification', 'Min Classification', 'Range Classification',
                 'Max Scan Angle Rank', 'Min Scan Angle Rank', 'Range Scan Angle Rank',
                 'Max User Data', 'Min User Data', 'Range User Data',
                 'Max Point Source ID', 'Min Point Source ID', 'Range Point Source ID',
                 'Max Red', 'Min Red', 'Range Red', 'Max Green', 'Min Green', 'Range Green',
                 'Max Blue', 'Min Blue', 'Range Blue']

    attribute_qc_file_name = csv_name

    # Open the file for append or create it
    with open(attribute_qc_file_name, 'a') as attr_file:
        attr_csv = csv.writer(attr_file)

        # Check if the file is empty
        attr_is_empty = os.stat(attribute_qc_file_name).st_size == 0

        # If it is empty write the file attribute names
        if attr_is_empty:
            attr_csv.writerow(attr_keys)

        # Compute Max, Min and Range for every attribute
        maxX = max(test.x)
        minX = min(test.x)
        rangeX = maxX - minX
        maxY = max(test.y)
        minY = min(test.y)
        rangeY = maxY - minY
        maxZ = max(test.z)
        minZ = min(test.z)
        rangeZ = maxZ - minZ
        maxIntensity = max(test.intensity)
        minIntensity = min(test.intensity)
        rangeIntensity = maxIntensity - minIntensity
        maxReturnNumber = max(test.return_num)
        minReturnNumber = min(test.return_num)
        rangeReturnNumber = maxReturnNumber - minReturnNumber
        maxNumberOfReturns = max(test.num_returns)
        minNumberOfReturns = min(test.num_returns)
        rangeNumberOfReturns = maxNumberOfReturns - minNumberOfReturns
        maxGps = max(test.gps_time)
        minGps = min(test.gps_time)
        rangeGps = maxGps - minGps

        # Convert gps_time into UTC time
        week_day_of_creation = date2gpswd(test.header.get_date().date())[0]
        maxDayCreation = gps2utc(week_day_of_creation, maxGps)
        minDayCreation = gps2utc(week_day_of_creation, minGps)
        rangeDayCreation = maxDayCreation - minDayCreation

        maxScanDirectionFlag = max(test.scan_dir_flag)
        minScanDirectionFlag = min(test.scan_dir_flag)
        rangeScanDirectionFlag = maxScanDirectionFlag - minScanDirectionFlag
        maxEdgeOfFlightLine = max(test.edge_flight_line)
        minEdgeOfFlightLine = min(test.edge_flight_line)
        rangeEdgeOfFlightLine = maxEdgeOfFlightLine - minEdgeOfFlightLine
        maxClassification = max(test.classification)
        minClassification = min(test.classification)
        rangeClassification = maxClassification - minClassification
        maxScanAngleRank = max(test.scan_angle_rank)
        minScanAngleRank = min(test.scan_angle_rank)
        rangeScanAngleRank = maxScanAngleRank - int(minScanAngleRank)
        maxUserData = max(test.user_data)
        minUserData = min(test.user_data)
        rangeUserData = maxUserData - minUserData
        maxPointSourceId = max(test.pt_src_id)
        minPointSourceId = min(test.pt_src_id)
        rangePointSourceId = maxPointSourceId - minPointSourceId
        maxRed = max(test.red)
        minRed = min(test.red)
        rangeRed = maxRed - minRed
        maxGreen = max(test.green)
        minGreen = min(test.green)
        rangeGreen = maxGreen - minGreen
        maxBlue = max(test.blue)
        minBlue = min(test.blue)
        rangeBlue = maxBlue - minBlue
        # TODO: maybe blue is not the final dimension

        # Make the row for the file
        attr_values = [las_file, maxX, minX, rangeX, maxY, minY, rangeY, maxZ, minZ, rangeZ,
                       maxIntensity, minIntensity, rangeIntensity,
                       maxReturnNumber, minReturnNumber, rangeReturnNumber,
                       maxNumberOfReturns, minNumberOfReturns, rangeNumberOfReturns,
                       maxGps, minGps, rangeGps,
                       maxDayCreation.strftime('%d-%B-%Y, %H:%M:%S'), minDayCreation.strftime('%d-%B-%Y, %H:%M:%S'),
                       str(rangeDayCreation),
                       maxScanDirectionFlag, minScanDirectionFlag, rangeScanDirectionFlag,
                       maxEdgeOfFlightLine, minEdgeOfFlightLine, rangeEdgeOfFlightLine,
                       maxClassification, minClassification, rangeClassification,
                       maxScanAngleRank, minScanAngleRank, rangeScanAngleRank,
                       maxUserData, minUserData, rangeUserData,
                       maxPointSourceId, minPointSourceId, rangePointSourceId,
                       maxRed, minRed, rangeRed, maxGreen, minGreen, rangeGreen, maxBlue, minBlue, rangeBlue]
        # Write the row to the file
        attr_csv.writerow(attr_values)


def header(test, las_file, csv_name='header.csv'):
    """
    Extract the header information from the file and write it in a CSV file.

    :param test: laspy object from which we get all header info and put it in a CSV file
    :type test: Laspy Object
    :param las_file: name of the laspy_object, i.e. 'Test.las'
    :type las_file: string
    :param csv_name: Name of the csv file (default 'header.csv')
    :type csv_name: string
    :return  create or appends to csv file that csv_name as name
    """

    # Required fields from the header of the file
    header_keys = ['File Name', 'File Signature', 'File Source Id', 'Global Encoding',
                   'Gps Time Type', 'Project Id', 'Version Major',
                   'Version Minor', 'System Identifier',
                   'Generating Software', 'File Creation Date', 'File Creation Day of Year', 'File Creation Year',
                   'Header Size', 'Offset to Point Data', 'Number of Variable Length Records',
                   'Point Data Format Id', 'Number of point records', 'Number of Points by Return Count',
                   'X scale factor', 'Y scale factor', 'Z scale factor',
                   'X offset', 'Y offset', 'Z offset',
                   'Max X', 'Min X', 'Range X', 'Max Y', 'Min Y', 'Range Y', 'Max Z', 'Min Z', 'Range Z']

    header_file_name = csv_name

    # Open the header csv file
    with open(header_file_name, 'a') as header_file:
        header_csv = csv.writer(header_file)

        # Check if it is empty
        header_is_empty = os.stat(header_file_name).st_size == 0

        # Write the header fields
        if header_is_empty:
            header_csv.writerow(header_keys)

        # Get the day of creation
        date_of_creation = test.header.get_date()

        # Write the row
        header_csv.writerow([las_file, test.header.file_signature, test.header.file_source_id,
                             test.header.global_encoding, test.header.gps_time_type, test.header.guid,
                             test.header.version_major, test.header.version_minor,
                             re.sub(r'[\x00]', r'', test.header.system_id),
                             re.sub(r'[\x00]', r'', test.header.software_id),
                             date_of_creation.strftime('%d-%B-%Y'), date_of_creation.strftime('%j'),
                             date_of_creation.year, test.header.header_size, test.header.data_offset,
                             len(test.header.vlrs), test.header.data_format_id, test.header.records_count,
                             test.header.point_return_count,
                             test.header.scale[0], test.header.scale[1], test.header.scale[2],
                             test.header.offset[0], test.header.offset[1], test.header.offset[2],
                             test.header.max[0], test.header.min[0], (test.header.max[0] - test.header.min[0]),
                             test.header.max[1], test.header.min[1], (test.header.max[1] - test.header.min[1]),
                             test.header.max[2], test.header.min[2], (test.header.max[2] - test.header.min[2])])


# Empty dict for the extra attributes
extra_dims_dict = defaultdict(list)


def extra_attr(test, las_file):
    """
    Extract the extra attributes and put them into the empty dict

    :param test: laspy object from which we get all header info and put it in a CSV file
    :type test: Laspy Object
    :param las_file: name of the laspy_object, i.e. 'Test.las'
    :type las_file: string
    :return  it is populating the empty dict
    """
    # Append the new name of the file
    extra_dims_dict['File Name'].append(las_file)

    # Check if the file has extra dimensions in the point_format lookup which is dictionary
    # for all the attributes plus the extra ones. I noticed that after the dimension blue
    # it is storing all the extra attributes.
    extra_dims = False
    has_extra_dims = False
    for key, val in test.point_format.lookup.items():
        if extra_dims:
            has_extra_dims = True
            break
        if key == 'blue':
            extra_dims = True

    extra_dims = False

    # Get the names of the extra attributes
    keys = [i for i in extra_dims_dict if extra_dims_dict[i] != extra_dims_dict.default_factory()]

    # Compute the Max Min and Range of the extra attributes
    if has_extra_dims:
        extra_dims = False
        has_extra_dims = False
        for key, val in test.point_format.lookup.items():
            if extra_dims:
                # If we have some files that don't have the same extra attr
                # as the previous one we put empty places in the csv file
                # so it would be complete
                if extra_dims_dict['File Name'].index(las_file) != len(extra_dims_dict['Max ' + key]):
                    for i in range(0, extra_dims_dict['File Name'].index(las_file)):
                        extra_dims_dict['Max ' + key].insert(0, '')
                        extra_dims_dict['Min ' + key].insert(0, '')
                        extra_dims_dict['Range ' + key].insert(0, '')
                extra_dims_dict['Max ' + key].append(max(test.reader.get_dimension(key)))
                extra_dims_dict['Min ' + key].append(min(test.reader.get_dimension(key)))
                extra_dims_dict['Range ' + key]. \
                    append(max(test.reader.get_dimension(key)) - min(test.reader.get_dimension(key)))
            if key == 'blue':
                extra_dims = True

        extra_dims = False

    limit = len(extra_dims_dict['File Name'])

    # Fill the empty column for different files
    for key in keys:
        column = extra_dims_dict[key]

        if len(column) != limit:
            for i in range(len(column), limit):
                column.append('')


def write_extra_dim(csv_name='extra_attributes.csv'):
    """
    Write the extra attributes to a CSV file. It is NOT appending to a file.

    :param csv_name: Name of the csv file.
    :type csv_name: string
    :return  It is making and writing the csv file
    """
    # Names of the extra dimensions
    field_names = [i for i in extra_dims_dict if extra_dims_dict[i] != extra_dims_dict.default_factory()]
    with open(csv_name, 'w', ) as header_csv:
        writer = csv.writer(header_csv)

        # Write the header of the csv file
        writer.writerow(field_names)

        # Write the rows of the csv file from the dictionary
        writer.writerows(zip(*[extra_dims_dict[key] for key in field_names]))


def extract_qc(location="./", location_qc="./", attribute_name='attribute.csv', header_name='header.csv',
               extra_attr_name='extra_attributes.csv'):
    """
    Listing a location on the system. From all the las files we extract the header information, the attribute
    MaxMinRange and the extra attributes into three csv files in location_qc folder.

    :param location: Location where to search for las files (default './')
    :param location_qc: Where to store the CSV files (default './')
    :param attribute_name: Name for attribute information (default 'attribute.csv')
    :param header_name: Name for header information (default 'header.csv')
    :param extra_attr_name: Name for extra attributes (default 'extra_attributes.csv')
    """

    # If the location for the qc is not there. Make it
    if not os.path.exists(location_qc):
        os.makedirs(location_qc)

    # List all the files specified under location param
    for las_file in os.listdir(location):
        # Check for las files
        filename, file_extension = os.path.splitext(las_file)
        if file_extension == '.las':
            print('Starting file: ' + filename)
            # Read the laspy object
            test = File(location+filename+'.las', mode='r')
            # Run attribute
            attribute(test, filename, csv_name=(location_qc + attribute_name))
            # Run header
            header(test, filename, csv_name=(location_qc + header_name))
            # Populate the empty dictionary
            extra_attr(test, filename)
            test.close()
    # Write the extra dimensions
    write_extra_dim(csv_name=(location_qc + extra_attr_name))


def voxel_count(infile, scale, offset=0.0):
    """
    Calculate the number of points in a voxel with a size of one side define with scale

    :param infile: Laspy object from where we get the coordinates
    :type infile: laspy object
    :param scale: size of one side of the voxel, so if you want 10cm voxels, scale = 0.1
    :type scale: float
    :param offset: Set offset for coords to be integer * scale + offset (default '0,0')'
    :type offset: float
    :return  array of number of points which each point has in it's voxel
    """
    # depending on whether you want a 3d or 2d voxel (e.g. x = inFile.x, y = inFile.y, z = inFile.z)
    # the voxels will have side coordinates which are integer multiples of scale
    # unless an offset is set, in which case they will be integer * scale + offset

    # turns coordinates into integers so we can use np.unique to count
    # u is the size of the voxel in meters (i.e. the side length)
    coords = np.stack((infile.x, infile.y, infile.z), axis=1)

    ar = np.floor((coords - offset) / scale).astype(int)
    unq, ind, inv, cnt = np.unique(ar, return_index=True, return_inverse=True, return_counts=True, axis=0)

    return cnt[inv]


def radial_count_v1(infile, radius):
    """
    Calculate the number of points which each point has in a circle with a radius specified.
    Fast for small files. Memory inefficient.

    :param infile: Laspy object from where we get the coordinates
    :type infile: laspy object
    :param radius: radius of the circle
    :type radius: float
    :return  array of number of points which each point has in it's circle
    """
    # radius is the radius of the circle or sphere  you want each point to use
    # coords is similar to above, again you can use 2d or 3d points
    coords = np.stack((infile.x, infile.y, infile.z), axis=1)
    lv = np.vectorize(len)  # this is so we can take lengths quickly, as discussed
    neigh = NearestNeighbors(radius=radius)
    neigh.fit(coords)
    d, i = neigh.radius_neighbors(coords)
    return lv(i)


def radial_count_v2(infile, radius):
    """
    Calculate the number of points which each point has in a circle with a radius specified.
    Memory efficient. Version 2 of radial_count. PROBABLY AVOID THIS!

    :param infile: Laspy object from where we get the coordinates
    :type infile: laspy object
    :param radius: radius of the circle
    :type radius: float
    :return  array of number of points which each point has in it's circle
    """

    coords = np.stack((infile.x, infile.y, infile.z), axis=1)
    done = np.zeros(coords.shape[-2], dtype=bool)
    ppm = np.empty(coords.shape[-2], dtype=int)
    k = 1
    while not (done.all()):
        nhbrs = NearestNeighbors(n_neighbors=k, algorithm="kd_tree").fit(coords)
        distances, indices = nhbrs.kneighbors(coords[~ done, :])  # (num_pts,k)
        k_found = distances[:, -1] >= radius
        not_done_save = ~done
        done[~done] = k_found
        ppm_not_done = ppm[not_done_save]
        ppm_not_done[k_found] = k
        ppm[not_done_save] = ppm_not_done
        k += 1
    return ppm


def radial_count_v3(infile, clip=0.5, classif=4, num_neighb=51):
    """
    Calculate the number of points which each point in number of points close to the point and their distance.
    The function now works for whole files regardless on the classif parameter.

    :param infile: Laspy object from where we get the coordinates
    :type infile: laspy object
    :param clip: radius of the circle
    :type clip: float
    :param classif: around which classification should the clip happen
    :type classif: int
    :param num_neighb: how much neighbours of points should we look at
    :type num_neighb: int
    :return  array of number of points which each point has in it's neighbours
    """

    cls = infile.Classification
    x_array = infile.x
    y_array = infile.y
    z_array = infile.z
    intensity = infile.Intensity
    # All classification 4 from the dataset
    class10 = (cls == classif)

    # XYZ dataset
    coords = np.vstack((x_array, y_array, z_array))

    # XYZ dataset with all classification 4
    # coords_flight = np.vstack((x_array[class10], y_array[class10], z_array[class10]))
    if len(coords[0]) != 0:

        # n_neighbors is the number of neighbors of a point, I can't take everything within 1m
        # but I can take a lot of them for now 11
        if len(coords[0]) < num_neighb:
            n_neigh = len(coords[0])
        else:
            n_neigh = num_neighb

        nhbrs = NearestNeighbors(n_neighbors=n_neigh, algorithm="kd_tree").fit(np.transpose(coords))
        distances, indices = nhbrs.kneighbors(np.transpose(coords))

        # axis = 0 along columns, axis  = 1 along rows
        num_neighbours = np.sum(distances < clip, axis=1, dtype=np.uint16)

        volume = (4 / 3) * np.pi * (clip ** 3)

        density = num_neighbours / volume

        intensity = num_neighbours

    return intensity


def radial_count_v4(infile, k=51, radius=0.5, spacetime=True, v_speed=2, N=20):
    if v_speed == 0:
        spacetime = False

    x_array = infile.x
    y_array = infile.y
    z_array = infile.z
    gps_time = infile.gps_time
    intensity = infile.Intensity
    classification = infile.Classification

    # create array of length x_array
    num_neighbours = np.zeros(x_array.size, dtype=np.uint16)

    if len(x_array) != 0:
        times = list([np.quantile(gps_time, q=float(i) / float(N)) for i in range(N + 1)])

        for i in range(N):
            time_range = (times[i] <= gps_time) * (gps_time <= times[i + 1])

            coords = np.vstack((x_array[time_range], y_array[time_range], z_array[time_range]) + spacetime * (
                v_speed * gps_time[time_range],))

            distances, indices = NearestNeighbors(n_neighbors=k, algorithm="kd_tree").fit(
                np.transpose(coords)).kneighbors(np.transpose(coords))
            neighbours = (coords)[:, indices]

            num_neighbours[time_range] = np.sum(distances < radius, axis=1)

        intensity = num_neighbours

    return intensity


def getWKT_PRJ_file(filename, epsg_code):
    """
    Write the projection file.

    :param filename: name of the PRJ file
    :type filename: string
    :param epsg_code:
    :type epsg_code: string
    :return: .prj file
    """
    import urllib.request
    # access projection information
    wkt = urllib.request.urlopen("http://spatialreference.org/ref/epsg/{0}/prettywkt/".format(epsg_code))
    # remove spaces between charachters
    string_wkt = wkt.read().decode('utf-8')
    remove_spaces = string_wkt.replace(" ", "")
    # place all the text on one line
    output = remove_spaces.replace("\n", "")

    # create the .prj file
    prj = open(filename+".prj", "w")

    prj.write(output)
    prj.close()


def polygon_select(infile, resolution=10, classif=15, classed='polygon'):
    """
    Make a polygon file which is going to give the vegetation on maps (json file) or google earth (ESRI shape file).

    :param infile: object on which to extract the vegetation
    :type infile: laspy object
    :param resolution: multiply the coords by this number to get better images
    :type resolution: float
    :param classif: class number for which to extract the polygons (default 15 (Vegetation))
    :type classif: int
    :param classed: it tells are we going to output polygon or points position
    :type classed: string
    :return: json, shp, xlsx, csv and prj files
    """
    # Reading a las file from the current location
    inFile = infile

    # Resolution of the image from the dataset
    res = resolution

    # xyz dataset multiply by the resolution
    xyz = np.dstack((inFile.x * res, inFile.y * res, inFile.z * res))[0]

    # Mask of the vegetation
    mask15 = inFile.Classification == classif
    xyz = xyz[mask15]

    # Max and min of X, Y and Z for the image
    xmax = max(inFile.x * res)
    xmin = min(inFile.x * res)

    ymax = max(inFile.y * res)
    ymin = min(inFile.y * res)

    zmax = max(inFile.z * res)
    zmin = min(inFile.z * res)

    # Resolution of the image
    shape = (int(xmax - xmin + 1), int(ymax - ymin + 1))

    # Array of the image X by Y
    img = np.ma.array(np.ones(shape) * (zmax + 1))

    # Populate the pixels
    for inp in xyz:
        img[int(inp[0] - xmin), int(inp[1] - ymin)] = int(inp[2])

    # Set up the mask
    img.mask = (img == zmax + 1)

    # Save it with skimage (Yet to figure out why!)
    io.imsave('foo.png', img)
    # imageio.imwrite('foo.png', img)

    # Read it with OpenCV (More libraries then skimage)
    im_in = cv2.imread('./foo.png')

    # Make a grayscale image
    gray = cv2.cvtColor(im_in, cv2.COLOR_BGR2GRAY)

    # Compute the edges with canny algorithm
    edged = cv2.Canny(gray, 30, 200)
    cv2.imwrite('edged1.png', edged)

    # Find the contours of the image (It is going to be every hole)
    contours, hierarchy = cv2.findContours(edged, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_NONE)
    # Draw them on the image
    cv2.drawContours(im_in, contours, -1, (0, 255, 0), 3)

    # Fill the polygons
    cv2.fillPoly(im_in, contours, color=(255, 255, 255))
    cv2.imwrite("filledPoly.png", im_in)

    # Threshold.
    # Set values equal to or above 220 to 0.
    # Set values below 220 to 255.
    th, im_th = cv2.threshold(im_in, 220, 255, cv2.THRESH_BINARY_INV)

    # Copy the thresholded image.
    im_floodfill = im_th.copy()

    # Mask used to flood filling.
    # Notice the size needs to be 2 pixels than the image.
    h, w = im_th.shape[:2]
    mask = np.zeros((h + 2, w + 2), np.uint8)

    # Floodfill from point (0, 0)
    cv2.floodFill(im_floodfill, mask, (0, 0), 255)

    # Invert floodfilled image
    im_floodfill_inv = cv2.bitwise_not(im_floodfill)

    # Combine the two images to get the foreground.
    im_out = im_th | im_floodfill_inv

    cv2.imwrite('filled2.png', im_out)
    image = cv2.imread('./filled2.png')

    # Again from the filled picture make a grayscale
    gray = cv2.cvtColor(image, cv2.COLOR_RGB2GRAY)

    # Compute the edges with Canny on the filled image
    edged = cv2.Canny(gray, 30, 200)

    # Close down the contours with adding some pixels
    kernel = cv2.getStructuringElement(cv2.MORPH_ELLIPSE, (3, 3))
    dilated = cv2.dilate(edged, kernel)
    cv2.imwrite('edged.png', dilated)

    # Again compute the contours on the filled image
    contours, hierarchy = cv2.findContours(dilated, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
    cv2.drawContours(image, contours, -1, (0, 255, 0), -1)

    # We need to make a "bad" json file for our MAP plotting into javascript google maps
    # and we do that with a simple string in which we apply the array of the (x,y) points
    # for the polygons
    jsonFile = "data = '{\"Polygons\":[\\\n"

    # The number of the polygon
    plid = 1

    # Full polygons of the shape file
    polygons = []
    areas = []
    centroids = []
    # Converting projected coordinates to lat-lon
    # https://gis.stackexchange.com/questions/78838/converting-projected-coordinates-to-lat-lon-using-python
    inProj = Proj(init='epsg:27700')
    outProj = Proj(init='epsg:4326')

    # csv name
    csvName = infile.filename.split('.')[0]+'_Classification'+str(classif)+'_'+classed+'.csv'

    # We also need the (x,y) points the area and the centroid point in ASCII
    with open(csvName, 'w') as csvFile:
        writer = csv.writer(csvFile)

        # Headers for the CSV file
        writer.writerow(['PlID', 'PtX', 'PtY', 'cX', 'cY', 'Area'])
        pId = []
        cntX = []
        cntY = []
        # For each polygon (contour) we write it in the JSON string and in the file
        for cnt in contours:

            jsonFile += '['

            # polygon id
            pId.append(plid)

            # OpenCV moments which get us to compute the centroid point
            M = cv2.moments(cnt)
            if M['m00'] != 0:
                cx = int(M['m10'] / M['m00'])
                cy = int(M['m01'] / M['m00'])

            # Area of the polygon
            area = cv2.contourArea(cnt)/(resolution * resolution)

            # Perimeter of the polygon (Just in case)
            perimeter = cv2.arcLength(cnt, True)

            # Centroid point on the image
            cv2.circle(image, (cx, cy), 3, (0, 0, 255), -1)

            # Each shape polygon separate
            polygon = []
            centroid = []

            # For each point in the polygon we must compute the real x,y point
            # from the image pixels and write them in the CSV and JSON
            for pts in cnt:
                row = [plid, (pts[0][1] / 10 + xmin / 10), (pts[0][0] / 10 + ymin / 10), (cy / 10 + xmin / 10),
                       (cx / 10 + ymin / 10), area]

                # convert the projected coordinates into lat-lon
                coord1, coord2 = transform(inProj, outProj, (pts[0][1] / 10 + xmin / 10), (pts[0][0] / 10 + ymin / 10))
                polygon.append((coord1, coord2))

                jsonFile += "{\"lat\":" + str(row[1]) + ",\"lng\":" + str(row[2]) + "},\\\n"
                writer.writerow(row)

            writer.writerow(
                [plid, (cnt[0][0][1] / 10 + xmin / 10), (cnt[0][0][0] / 10 + ymin / 10), (cy / 10 + xmin / 10),
                 (cx / 10 + ymin / 10), area])

            # We don't need comma in the final polygon in the "bad" JSON
            if plid == len(contours):
                jsonFile += "{\"lat\":" + str((cnt[0][0][1] / 10 + xmin / 10)) + ",\"lng\":" + str(
                    (cnt[0][0][0] / 10 + ymin / 10)) + "}]\\\n"
            else:
                jsonFile += "{\"lat\":" + str((cnt[0][0][1] / 10 + xmin / 10)) + ",\"lng\":" + str(
                    (cnt[0][0][0] / 10 + ymin / 10)) + "}],\\\n"

            plid += 1

            # Append to the whole shape polygons dataset
            polygons.append(polygon)

            # Append centroid points for each polygon
            cnt1, cnt2 = transform(inProj, outProj, (cy / 10 + xmin / 10), (cx / 10 + ymin / 10))
            centroid.append([cnt1, cnt2])

            # Excel coords
            cntX.append(cnt1)
            cntY.append(cnt2)

            # Append area of each polygon
            areas.append(area)

            centroids.append(centroid)

    # print(polygons)

    # Make xlsx file
    xlsx_polygon(infile.filename.split('.')[0]+'_Classification'+str(classif)+'_'+classed, np.array([pId, cntX, cntY, areas]))

    # Description for the shape file
    schema = {
        'geometry': 'Polygon',
        'properties': {'class': 'str',
                       'tileNo': 'str',
                       'polygonNo': 'int',
                       'area(m2)': 'float',
                       'centroidXY': 'str',
                       'desc': 'str',
                       'voltage': 'str',
                       'srcName': 'str',
                       }
    }

    if classed == 'points':
        schema = {
            'geometry': 'Polygon',
            'properties': {'class': 'str',
                           'tileNo': 'str',
                           'poleNo': 'int',
                           }
        }

    # Name of the shape file
    nameSHP = infile.filename.split('.')[0]+'_Classification'+str(classif)+'_'+classed+'.shp'

    # Write the shapefile
    with fiona.open(nameSHP, 'w', 'ESRI Shapefile', schema) as c:
        for poly in range(len(polygons)):
            if classed == 'polygon':
                plg = Polygon(polygons[poly])

                centroidsPnt = str(str(centroids[poly][0][0])+' '+str(centroids[poly][0][1]))
                # print(areas[poly])
                c.write({
                    'geometry': mapping(plg),
                    'properties': {'class': 'Vegetation 3m',
                                   'tileNo': infile.filename.split('.')[0],
                                   'polygonNo': poly+1,
                                   'area(m2)': areas[poly],
                                   'centroidXY': centroidsPnt,
                                   'desc': '',
                                   'voltage': '',
                                   'srcName': ''
                                   },
                })
            elif classed == 'points':
                plg = Point(centroids[poly][0][0], centroids[poly][0][1])

                c.write({
                    'geometry': mapping(plg.buffer(0.00001)),
                    'properties': {'class': 'Pole',
                                   'tileNo': infile.filename.split('.')[0],
                                   'poleNo': poly + 1
                                   },
                })

    # We close down the "bad" JSON
    jsonFile += "]}';"
    # Write it in a .json file
    with open(infile.filename.split('.')[0]+'_Classification'+str(classif)+'_'+classed+".json", "w") as text_file:
        text_file.write(jsonFile)
    csvFile.close()
    text_file.close()

    # Get the projection file
    getWKT_PRJ_file(infile.filename.split('.')[0]+'_Classification'+str(classif)+'_'+classed, "4326")

    # Delete all the helping images
    os.system("rm ./*.png")
    # Save the last output
    cv2.imwrite('im_out.png', image)


def xlsx_polygon(filename, data):
    """
    Make a polygon file which is going to give the vegetation on maps (json file) or google earth (ESRI shape file).

    :param filename: name of the file which will be put on the spreadsheet
    :type filename: string
    :param data: array of columns for the xlsx data
    :type data: ndarray
    :return: writes a xlsx file to the system
    """

    workbook = xlsxwriter.Workbook(filename.split('.')[0]+'.xlsx')
    # By default worksheet names in the spreadsheet will be
    # Sheet1, Sheet2 etc., but we can also specify a name.
    worksheet = workbook.add_worksheet("My sheet")

    # Start from the first cell. Rows and
    # columns are zero indexed.
    row = 0
    col = 0

    # Header of the excel file
    worksheet.write_row(row, col, ['Polygon ID', 'Centroid X', 'Centroid Y', 'Area(m2)'])
    row += 1

    # Transpose the data each row to be in a different array
    data = data.T

    # Write the data with specific column width
    for item in data:
        worksheet.set_column(row, col, len(str(item))/2)
        worksheet.write_row(row, col, item)
        row += 1

    workbook.close()


def column_string(n):
    """
    Turn number into column letter.

    :param n: number of the column
    :return: string
    """
    string = ""
    while n > 0:
        n, remainder = divmod(n - 1, 26)
        string = chr(65 + remainder) + string
    return string


def fill_shape_file(filename_poly, filename_lines):
    """
    Fill the description of the polygon file that is near the lines.

    :param filename_poly: name of the shape file with the polygons
    :type filename_poly: string
    :param filename_lines: name of the shape file with the lines
    :type filename_lines: string
    :return:
    """
    import geopandas as gpd
    from shapely.wkt import loads

    # Get the shape file
    sf_poly = shapefile.Reader(filename_poly)
    print('number of shapes imported:', len(sf_poly.shapes()))
    print('geometry attributes in each shape:')

    # dict for filled polygons
    poly_filled = []

    # open the existing polygon without the extra attr
    with fiona.open(filename_poly+'.shp') as polygons:

        # copy the schema from the shape file
        schema = polygons.schema.copy()

        # iterate through each polygon
        for pol in polygons:

            # create polygon string with cooordinates
            poly_string = 'POLYGON (('

            indx = 1

            # get the coordinates into projection 27700 same as the shape conductor file
            for coord in pol['geometry']['coordinates'][0]:
                inProj = Proj(init='epsg:4326')
                outProj = Proj(init='epsg:27700')
                cnt1, cnt2 = transform(inProj, outProj, coord[0], coord[1])
                if indx != len(pol['geometry']['coordinates'][0]):
                    poly_string += '' + str(cnt1) + ' ' + str(cnt2) + ', '
                else:
                    poly_string += '' + str(cnt1) + ' ' + str(cnt2) + ' '
                indx += 1
            # close the polygon
            poly_string += '))'

            min_distance = sys.maxsize
            with fiona.open(filename_lines+'.shp') as lines:

                inProj = Proj(init='epsg:27700')
                outProj = Proj(init='epsg:4326')

                line_string = 'LINESTRING ('

                ind = 0
                min_ind = ind
                for line in lines:
                    indx = 1
                    line_string = 'LINESTRING ('
                    for coord in line['geometry']['coordinates']:
                        # cnt1, cnt2 = transform(inProj, outProj, coord[0], coord[1])
                        # print(cnt1, cnt2)
                        if indx != len(line['geometry']['coordinates']):
                            line_string += ''+str(coord[0])+' '+str(coord[1])+', '
                        else:
                            line_string += '' + str(coord[0]) + ' ' + str(coord[1]) + ' '
                        indx += 1
                    line_string += ')'

                    lline = loads(line_string)
                    g_line = gpd.GeoSeries(lline)

                    poly = loads(poly_string)
                    g_poly = gpd.GeoSeries(poly)
                    if g_poly.distance(g_line)[0] < min_distance:
                        min_distance = g_poly.distance(g_line)[0]
                        min_ind = ind
                    ind += 1

                pol['properties']['desc'] = lines[min_ind]['properties']['desc']
                pol['properties']['voltage'] = lines[min_ind]['properties']['voltage']
                pol['properties']['srcName'] = lines[min_ind]['properties']['srcName']

                poly_filled.append(pol)

                # print(pol)
                # print(min_distance)
                # print(min_ind)

    extra_data = []
    with fiona.open(filename_poly+'.shp', 'w', 'ESRI Shapefile', schema) as c:
        for elem in poly_filled:
            c.write({'properties': elem['properties'], 'geometry': mapping(shape(elem['geometry']))})
            extra_data.append([elem['properties']['desc'], elem['properties']['voltage'], elem['properties']['srcName']])

    wb = load_workbook(filename_poly+'.xlsx')

    sheet = wb.active

    max_col = sheet.max_column
    max_row = sheet.max_row

    sheet.cell(row=1, column=max_col+1).value = 'Description'
    sheet.cell(row=1, column=max_col + 2).value = 'Voltage'
    sheet.cell(row=1, column=max_col + 3).value = 'Source Name'

    sheet.column_dimensions[column_string(max_col + 1)].width = 100
    sheet.column_dimensions[column_string(max_col + 2)].width = sheet.column_dimensions['A'].width
    sheet.column_dimensions[column_string(max_col + 3)].width = sheet.column_dimensions['A'].width

    for i in range(2, max_row+1):
        sheet.cell(row=i, column=max_col+1).value = extra_data[i-2][0]
        sheet.cell(row=i, column=max_col + 2).value = extra_data[i-2][1]
        sheet.cell(row=i, column=max_col + 3).value = extra_data[i-2][2]

    wb.save(filename_poly+'.xlsx')
